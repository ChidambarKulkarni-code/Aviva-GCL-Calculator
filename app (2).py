import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook

# --------------------------------------------------
# File paths — place both Excel files in the same folder as app.py
# --------------------------------------------------
SINGLE_LIFE_FILE = Path("Aviva_Final_-_Lap___HL-_single_life.xlsx")
JOINT_LIFE_FILE  = Path("AVIVA_GCL-_joint_life.xlsx")

# Sheet name lookup: (product, life_type) → (file_path, sheet_name)
SHEET_MAP = {
    ("Home Loan",            "Single Life"): (SINGLE_LIFE_FILE, "Home Loan"),
    ("Loan Against Property","Single Life"): (SINGLE_LIFE_FILE, "Lap"),
    ("Home Loan",            "Joint Life"):  (JOINT_LIFE_FILE,  "Homeloan"),
    ("Loan Against Property","Joint Life"):  (JOINT_LIFE_FILE,  "Lap"),
}

REQUIRED_COLUMNS = [
    "Customer Name",
    "Primary Age",
    "Secondary Age",       # Required for Joint Life; may be blank for Single Life
    "Loan Amount",
    "Tenure Years",        # Tenure in YEARS (the rate tables use yearly tenures)
]

# --------------------------------------------------
# Page config
# --------------------------------------------------
st.set_page_config(
    page_title="Aviva GCL Premium Calculator",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("Aviva Group Credit Life (GCL) Premium Calculator")
st.caption(
    "Portfolio-level automated premium computation for Single and Joint Life credit insurance. "
    "Rate tables are read directly from the Aviva rate Excel files."
)
st.divider()

# --------------------------------------------------
# Rate table loader
# --------------------------------------------------
@st.cache_data
def load_rate_table(file_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Reads a rate sheet from an Aviva Excel file.
    Returns a DataFrame with 'Entry Age' as the first column and
    integer-year tenure values as the remaining columns.
    """
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the header row — it contains "AGE" in the first cell
    header_idx = None
    for i, row in enumerate(all_rows):
        cell = str(row[0]).strip().upper() if row[0] is not None else ""
        if "AGE" in cell:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"Could not find 'AGE' header row in sheet '{sheet_name}' of '{file_path}'")

    header_row = all_rows[header_idx]
    # Tenures are in columns 1+ (skip the age label column)
    tenures = []
    for v in header_row[1:]:
        if v is not None:
            try:
                tenures.append(int(float(str(v))))
            except (ValueError, TypeError):
                pass

    records = []
    for row in all_rows[header_idx + 1:]:
        if row[0] is None:
            continue
        try:
            age = int(float(str(row[0])))
        except (ValueError, TypeError):
            continue
        record = {"Entry Age": age}
        for t, v in zip(tenures, row[1:len(tenures) + 1]):
            try:
                record[t] = float(v) if v is not None else np.nan
            except (ValueError, TypeError):
                record[t] = np.nan
        records.append(record)

    df = pd.DataFrame(records)
    return df


def get_rate(rate_table: pd.DataFrame, age: int, tenure_years: int):
    """
    Look up the premium rate for a given age and tenure (in years).
    Returns (rate, remark).
    """
    tenure_cols = [c for c in rate_table.columns if c != "Entry Age"]

    if tenure_years not in tenure_cols:
        available = sorted(tenure_cols)
        return None, f"Tenure {tenure_years}y not in table (available: {available[0]}–{available[-1]}y)"

    match = rate_table[rate_table["Entry Age"] == age]
    if match.empty:
        min_age = int(rate_table["Entry Age"].min())
        max_age = int(rate_table["Entry Age"].max())
        return None, f"Age {age} not in table (range: {min_age}–{max_age})"

    rate = match[tenure_years].iloc[0]
    if pd.isna(rate):
        return None, f"Rate is blank for age {age}, tenure {tenure_years}y"

    return float(rate), "OK"


def calculate_premium(row, rate_table: pd.DataFrame, life_type: str):
    """
    Compute the premium for one customer row.
    The rate tables store the premium per lakh of loan amount.
    Premium = (Loan Amount / 1,00,000) × Rate
    """
    try:
        age_1       = int(float(row["Primary Age"]))
        loan_amount = float(row["Loan Amount"])
        tenure_yrs  = int(float(row["Tenure Years"]))
    except Exception:
        return pd.Series([np.nan, np.nan, "Invalid numeric input"])

    rate_1, remark_1 = get_rate(rate_table, age_1, tenure_yrs)
    if rate_1 is None:
        return pd.Series([np.nan, np.nan, f"Primary: {remark_1}"])

    if life_type == "Joint Life":
        sec_age_raw = row.get("Secondary Age", "")
        if pd.isna(sec_age_raw) or str(sec_age_raw).strip() in ("", "0"):
            return pd.Series([np.nan, np.nan, "Missing Secondary Age for Joint Life"])
        try:
            age_2 = int(float(sec_age_raw))
        except (ValueError, TypeError):
            return pd.Series([np.nan, np.nan, "Invalid Secondary Age"])

        rate_2, remark_2 = get_rate(rate_table, age_2, tenure_yrs)
        if rate_2 is None:
            return pd.Series([np.nan, np.nan, f"Secondary: {remark_2}"])

        # Joint rate table already encodes the joint premium directly
        # The table is indexed by the *primary* borrower's age.
        # Use the higher age as the pricing age (standard actuarial practice)
        pricing_age = max(age_1, age_2)
        rate_final, remark_final = get_rate(rate_table, pricing_age, tenure_yrs)
        if rate_final is None:
            return pd.Series([np.nan, np.nan, f"Joint pricing age {pricing_age}: {remark_final}"])
        remark = f"Joint (pricing age={pricing_age})"
    else:
        rate_final = rate_1
        remark = "Single Life"

    premium = (loan_amount / 100_000) * rate_final
    return pd.Series([round(rate_final, 5), round(premium, 2), remark])


def to_excel_download(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Premium Output")
    output.seek(0)
    return output.read()


def read_customer_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file)


# --------------------------------------------------
# Rate file availability check
# --------------------------------------------------
missing_files = [f for f in [SINGLE_LIFE_FILE, JOINT_LIFE_FILE] if not f.exists()]
if missing_files:
    st.error(
        "**Rate file(s) missing.** Place the following files in the same folder as `app.py`:\n\n"
        + "\n".join(f"- `{f.name}`" for f in missing_files)
    )
    st.stop()

# --------------------------------------------------
# Sidebar — configuration
# --------------------------------------------------
st.sidebar.title("⚙️ Underwriting Controls")

product = st.sidebar.selectbox(
    "Product Type",
    ["Home Loan", "Loan Against Property"],
    help="Select the credit product being insured."
)

life_type = st.sidebar.selectbox(
    "Life Structure",
    ["Single Life", "Joint Life"],
    help="Single Life covers one borrower; Joint Life covers two."
)

# Resolve which file + sheet to use
sheet_key = (product, life_type)
if sheet_key not in SHEET_MAP:
    st.sidebar.error("Combination not supported.")
    st.stop()

rate_file_path, rate_sheet_name = SHEET_MAP[sheet_key]
st.sidebar.success(f"📄 Rate file: `{rate_file_path.name}`\n\n📋 Sheet: `{rate_sheet_name}`")

# Load & preview rate table
try:
    rate_table = load_rate_table(rate_file_path, rate_sheet_name)
except Exception as e:
    st.sidebar.error(f"Failed to load rate table: {e}")
    st.stop()

tenure_range = sorted([c for c in rate_table.columns if c != "Entry Age"])
age_range    = sorted(rate_table["Entry Age"].tolist())

st.sidebar.markdown(
    f"**Age range:** {age_range[0]}–{age_range[-1]}  \n"
    f"**Tenure range:** {tenure_range[0]}–{tenure_range[-1]} years"
)

with st.sidebar.expander("👁 Preview Rate Table"):
    st.dataframe(rate_table.set_index("Entry Age"), use_container_width=True)

# --------------------------------------------------
# Main layout
# --------------------------------------------------
left_col, right_col = st.columns([2, 1])

with left_col:
    st.subheader("📂 Upload Customer Portfolio")
    uploaded_file = st.file_uploader(
        "Supported formats: .xlsx, .xls, .csv",
        type=["xlsx", "xls", "csv"]
    )

with right_col:
    st.subheader("Required Column Headers")
    st.markdown(
        "Your file must contain **exactly** these column names "
        "(extra columns are preserved in the output):"
    )
    for col in REQUIRED_COLUMNS:
        required = "" if col == "Secondary Age" else ""
        note = " *(blank for Single Life)*" if col == "Secondary Age" else ""
        note = " *(in years, e.g. 5)*" if col == "Tenure Years" else note
        st.markdown(f"- `{col}`{note} {required}")
    st.info(
        "**Tenure Years** must match the tenures in the rate table "
        f"({tenure_range[0]}–{tenure_range[-1]} years for the current selection)."
    )

# --------------------------------------------------
# Computation engine
# --------------------------------------------------
if uploaded_file is not None:
    try:
        customer_df = read_customer_file(uploaded_file)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

    st.divider()
    st.subheader("📋 Input Preview")
    st.dataframe(customer_df.head(20), use_container_width=True)
    st.caption(f"{len(customer_df):,} rows loaded")

    # Column validation
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in customer_df.columns]
    if missing_cols:
        st.error(f"Missing required columns: **{', '.join(missing_cols)}**")
        st.stop()

    st.divider()
    st.subheader("🧮 Computing Premiums…")

    with st.spinner("Processing portfolio…"):
        results = customer_df.apply(
            lambda row: calculate_premium(row, rate_table, life_type),
            axis=1
        )
        results.columns = ["Applied Rate (per Lakh ₹)", "Computed Premium (₹)", "Remark"]
        output_df = pd.concat([customer_df, results], axis=1)

    st.success("Done!")
    st.dataframe(output_df, use_container_width=True)

    # Summary metrics
    valid = output_df["Computed Premium (₹)"].dropna()
    errors = output_df[output_df["Remark"] != "Single Life"][output_df["Remark"] != "Joint Life"]  # noqa

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Records",         f"{len(output_df):,}")
    col2.metric("Successfully Computed", f"{len(valid):,}")
    col3.metric("Errors / Skipped",      f"{len(output_df) - len(valid):,}")
    col4.metric("Total Premium (₹)",     f"₹{valid.sum():,.2f}")

    # Error detail
    error_rows = output_df[output_df["Computed Premium (₹)"].isna()]
    if not error_rows.empty:
        with st.expander(f"⚠️ {len(error_rows)} row(s) with errors"):
            st.dataframe(
                error_rows[["Customer Name", "Primary Age", "Secondary Age",
                             "Loan Amount", "Tenure Years", "Remark"]],
                use_container_width=True
            )

    st.divider()

    # Download
    excel_bytes = to_excel_download(output_df)
    st.download_button(
        label="📥 Download Premium Output (.xlsx)",
        data=excel_bytes,
        file_name="Aviva_GCL_Premium_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
